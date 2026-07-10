import openpyxl, os, re, csv, glob

BASE = os.path.dirname(os.path.abspath(__file__))
RPT = os.path.join(BASE, 'reports')
MONTHS = {m: i+1 for i, m in enumerate(['January','February','March','April','May','June','July','August','September','October','November','December'])}

def month_key(fname):
    m = re.match(r'([A-Za-z]+) (\d{4})-', os.path.basename(fname))
    return int(m.group(2)), MONTHS[m.group(1)]

# ---- by Code: long panel (year, month, rsp, code, service, records, quantity, price, nrc_total, total)
rows = []
for f in glob.glob(os.path.join(RPT, '*-Monthly Telecom Report by Code.xlsx')):
    y, mo = month_key(f)
    wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    for r in ws.iter_rows(values_only=True):
        if r is None or len(r) < 10:
            continue
        # data rows: col B = RSP, col C = code, numeric total in col J
        if r[1] is not None and r[2] is not None and isinstance(r[9], (int, float)):
            rows.append([y, mo, r[1], str(r[2]), r[3], r[4], r[5], r[6], r[8], r[9]])
    wb.close()

rows.sort(key=lambda x: (x[0], x[1]))
with open(os.path.join(BASE, 'bycode_long.csv'), 'w', newline='', encoding='utf-8') as fh:
    w = csv.writer(fh)
    w.writerow(['year','month','rsp','code','service','records','quantity','price','nrc_total','total'])
    w.writerows(rows)
print('bycode rows:', len(rows))

# ---- Wireless: per-RSP per-tier subs
wrows = []
for f in glob.glob(os.path.join(RPT, '*-Wireless.xlsx')):
    y, mo = month_key(f)
    wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
    ws = None
    for sh in wb.worksheets:
        if hasattr(sh, 'iter_rows'):
            ws = sh
            break
    if ws is None:
        print('no data sheet in', f)
        wb.close()
        continue
    for r in ws.iter_rows(values_only=True):
        if r and len(r) >= 4 and r[0] is not None and r[1] is not None:
            wrows.append([y, mo, r[0], r[1], r[2], r[3]])
    wb.close()
wrows.sort(key=lambda x: (x[0], x[1]))
with open(os.path.join(BASE, 'wireless_rsp_tier.csv'), 'w', newline='', encoding='utf-8') as fh:
    w = csv.writer(fh)
    w.writerow(['year','month','rsp','tier','service','subs'])
    w.writerows(wrows)
print('wireless rows:', len(wrows))
